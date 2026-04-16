#!/usr/bin/env python3
"""Force Excel formula recalculation flag on workbook.

Opens an .xlsx file, marks all formula cells as needing recalculation,
and saves. The actual recalculation happens when opened in Excel/LibreOffice.

Usage:
    python3 scripts/recalc.py output.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not installed. Run: pip3 install --user openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Force Excel formula recalculation")
    parser.add_argument("file", help="Path to .xlsx file")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"❌ File not found: {path}")
        sys.exit(1)

    wb = load_workbook(path)

    # Set calcMode to auto and force recalc on open
    if wb.calculation is not None:
        wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

    wb.save(path)
    size_kb = path.stat().st_size / 1024
    print(f"✅ Recalc flagged: {path} ({formula_count} formulas, {size_kb:.1f} KB)")


if __name__ == "__main__":
    main()

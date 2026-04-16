#!/usr/bin/env python3
"""Generate a professional Excel (.xlsx) spreadsheet from JSON data.

Usage:
    python3 gen_xlsx.py --input data.json --output report.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

STYLES = {
    "corporate": {
        "header_fill": "1A365D",
        "header_font_color": "FFFFFF",
        "accent_fill": "EBF5FB",
        "border_color": "B0BEC5",
        "title_font_color": "1A365D",
    },
    "academic": {
        "header_fill": "744210",
        "header_font_color": "FFFFFF",
        "accent_fill": "FFF8E1",
        "border_color": "D7CCC8",
        "title_font_color": "744210",
    },
    "minimal": {
        "header_fill": "059669",
        "header_font_color": "FFFFFF",
        "accent_fill": "F0FDF4",
        "border_color": "E5E7EB",
        "title_font_color": "059669",
    },
}


def apply_sheet_style(ws, sheet_data, style):
    """Apply data and formatting to a worksheet."""
    title = sheet_data.get("title", "Sheet")
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    formulas = sheet_data.get("formulas", {})
    col_widths = sheet_data.get("column_widths", {})

    thin_border = Border(
        left=Side(style="thin", color=style["border_color"]),
        right=Side(style="thin", color=style["border_color"]),
        top=Side(style="thin", color=style["border_color"]),
        bottom=Side(style="thin", color=style["border_color"]),
    )

    header_font = Font(name="Calibri", size=11, bold=True, color=style["header_font_color"])
    header_fill = PatternFill(start_color=style["header_fill"],
                              end_color=style["header_fill"], fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    accent_fill = PatternFill(start_color=style["accent_fill"],
                              end_color=style["accent_fill"], fill_type="solid")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=style["title_font_color"])
    title_cell.alignment = Alignment(horizontal="center")

    # Header row (row 3, leaving row 2 blank)
    header_row = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows
    for r_idx, row in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Stripe alternate rows
            if r_idx % 2 == 1:
                cell.fill = accent_fill

    # Apply formulas (e.g., {"D5": "=SUM(D3:D4)"})
    for cell_ref, formula in formulas.items():
        ws[cell_ref] = formula

    # Column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Auto-width fallback
    if not col_widths and headers:
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_xlsx(data, style_name):
    """Generate an Excel workbook from structured JSON data."""
    style = STYLES.get(style_name, STYLES["corporate"])
    wb = Workbook()

    sheets = data.get("sheets", [])
    if not sheets:
        # Single-sheet mode: treat top-level data as one sheet
        sheets = [data]

    for i, sheet_data in enumerate(sheets):
        if i == 0:
            ws = wb.active
            ws.title = sheet_data.get("title", "Sheet1")[:31]
        else:
            ws = wb.create_sheet(title=sheet_data.get("title", f"Sheet{i+1}")[:31])
        apply_sheet_style(ws, sheet_data, style)

    return wb


def main():
    parser = argparse.ArgumentParser(
        description="Generate professional Excel (.xlsx) spreadsheet from JSON data")
    parser.add_argument("--input", required=True, help="Path to JSON file with data")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    parser.add_argument("--style", choices=list(STYLES.keys()), default="corporate",
                        help="Spreadsheet style (default: corporate)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = generate_xlsx(data, args.style)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    size_kb = output_path.stat().st_size / 1024
    sheet_count = len(wb.sheetnames)
    print(f"✅ Saved: {output_path} ({size_kb:.1f} KB, {sheet_count} sheet(s), style: {args.style})")


if __name__ == "__main__":
    main()
